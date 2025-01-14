import openpyxl
from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.conf import settings
from django.http import HttpResponse
from django.dispatch import receiver
from django.db.models.signals import post_save
from apps.usuarios.models import Aprendiz, Usuario
from django.core.mail import send_mail
from .forms import CustomLoginForm, CargarAprendicesForm
from django.core.exceptions import ValidationError
from datetime import datetime, timezone
from django.utils.timezone import now, timedelta
import logging
from openpyxl import load_workbook


def login_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')

    if request.method == 'POST':
        form = CustomLoginForm(data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(request, username=username, password=password)
            if user:
                login(request, user)
                return redirect('dashboard')
        return render(request, 'auth/login.html', {'form': form, 'error': 'Credenciales inválidas'})
    else:
        form = CustomLoginForm()
    return render(request, 'auth/login.html', {'form': form})


@login_required
def logout_view(request):
    logout(request)
    return redirect('login')


@login_required
def dashboard_view(request):
    if request.user.rol == 'coordinador':
        return redirect('coordinador_dashboard')
    elif request.user.rol == 'asesor':
        return redirect('asesor_dashboard')
    elif request.user.rol == 'aprendiz':
        return redirect('aprendiz_dashboard')
    else:
        return HttpResponse("No tienes un rol asignado.", status=403)


@login_required
def coordinador_dashboard(request):
    if request.user.rol == 'coordinador':
        return render(request, 'dashboards/coordinador_dashboard.html', {'user': request.user})
    else:
        return HttpResponse("Acceso denegado.", status=403)


@login_required
def asesor_dashboard(request):
    if request.user.rol == 'asesor':
        return render(request, 'dashboards/asesor_dashboard.html', {'user': request.user})
    else:
        return HttpResponse("Acceso denegado.", status=403)


@login_required
def aprendiz_dashboard(request):
    if request.user.rol == 'aprendiz':
        return render(request, 'dashboards/aprendiz_dashboard.html', {'user': request.user})
    else:
        return HttpResponse("Acceso denegado.", status=403)


#TODO terminar la funcionalidad del envio de correos
def cargar_aprendices(request):
    mensaje = None
    error = None

    if request.method == 'POST' and request.FILES.get('file'):
        archivo = request.FILES['file']
        registros_procesados = 0
        registros_omitidos = 0

        try:
            #cargar el archivo excel
            wb = load_workbook(archivo)
            sheet = wb.active

            # iterar sobre las filas del archivo excel
            for row in sheet.iter_rows(min_row=2, values_only=True):  #saltar la primera fila de encabezado
                nombre, apellido, correo = row

                try:
                    if correo and not Aprendiz.objects.filter(correo=correo).exists():
                        usuario = Usuario.objects.create(
                            username=correo,
                            email=correo,
                            first_name=nombre,
                            last_name=apellido,
                            password=Usuario.objects.make_random_password()
                        )
                        usuario.set_password(usuario.password)
                        usuario.save()

                        aprendiz = Aprendiz.objects.create(
                            nombre=nombre,
                            apellido=apellido,
                            correo=correo,
                            estado='nuevo'
                        )

                        # enviar correo si el estado es nuevo
                        """
                        if aprendiz.estado == 'nuevo':
                            send_mail(
                                'Bienvenido al Sistema de gestión de aprendices',
                                'Por favor, cambie su clave para poder acceder.',
                                settings.DEFAULT_FROM_EMAIL,
                                [correo],
                                fail_silently=False,
                            )
                        registros_procesados += 1
                        """
                    else:
                        registros_omitidos += 1
                except Exception as registro_error:
                    registros_omitidos += 1
                    print(f"Error procesando el registro {correo}: {registro_error}")
            mensaje = (
                f"Proceso completado. Registros procesados: {registros_procesados}. "
                f"Registros omitidos: {registros_omitidos}."
            )
        except Exception as e:
            error = f"Error general: {str(e)}"

    return render(request, 'dashboards/aprendiz_dashboard.html', {'mensaje': mensaje, 'error': error})
